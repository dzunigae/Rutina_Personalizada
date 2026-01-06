from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox
import pandas as pd
import random
from pathlib import Path

# BUSCAR ASSETS

def encontrar_rutina():
    documentos = Path.home() / "OneDrive" / "Documentos"  # funciona en Windows moderno

    for ruta in documentos.rglob("Rutina de ejercicio"):
        if ruta.is_dir():
            return ruta

    raise FileNotFoundError("No se encontró la carpeta 'Rutina de ejercicio'")

# DEFINICIÓN DE OBJETOS

# 1. Definición objeto ejercicio
class ejercicio:
    def __init__(self,nombre,nivel,equipo,pagina,veces_realizado):
        self.nombre = nombre
        self.nivel = nivel
        self.equipo = equipo
        self.pagina = pagina
        self.veces_realizado = veces_realizado

# 2. Definición objeto musculo
class musculo:
    def __init__(self,nombre,lista_ejercicios,veces_ejercitado):
        self.nombre = nombre
        self.lista_ejercicios = lista_ejercicios
        self.veces_ejercitado = veces_ejercitado

# 3. Definición objeto Grupo muscular
class grupo_muscular:
    def __init__(self,nombre,lista_musculos,veces_ejercitado):
        self.nombre = nombre
        self.lista_musculos = lista_musculos
        self.veces_ejercitado = veces_ejercitado

# 4. Definición objeto rutina
class rutina:
    def __init__(self,index,lista_grupos_musculares):
        self.index = index
        self.lista_grupos_musculares = lista_grupos_musculares

# DEFINICIÓN DE VARIABLES GLOBALES
MODELO = encontrar_rutina() / "Modelo_produccion.xlsx"
EJERCICIOS_LIST = []
MUSCULOS_LIST = []
GRUPOS_MUSCULARES_LIST = []
RUTINAS_LIST = []
TUPLAS_GP_MUS_EJC_LIST = []
rutina_txt = encontrar_rutina() / "rutina.txt"

# DEFINICIÓN DE FUNCIONES

def rellenar_una_plaza(plaz,tgmei):
    for tupla_element in tgmei:
        # xx. Si se encuentra:
        if plaz[1] == tupla_element[3]:
            # xx. La plaza queda asociada al ejercicio en cuestión.
            plaz[2] = next(obj for obj in EJERCICIOS_LIST if obj.nombre == tupla_element[2])
            # xx. Se da por ejercitado el grupo muscular y se eliminan las tuplas de la lista que tengan ese grupo muscular.
            tgmei = [tupla for tupla in tgmei if tupla_element[0] not in tupla]
            # xx. También se eliminan de la estructura de datos de tuplas todas aquellas que repitan el ejercicio que acaba de salir.
            tgmei = [tupla for tupla in tgmei if tupla_element[2] not in tupla]
            # xx. Añadir la información de la tupla elegida a la variable global
            TUPLAS_GP_MUS_EJC_LIST.append(tupla_element)
            break
    return tgmei

# MAIN

if __name__ == "__main__":
    usar_equipo = messagebox.askyesno("Equipo","¿Deseas incluir ejercicios con equipo?")

    # Construir el df necesario del modelo
    MODELO_DF = pd.read_excel(MODELO)

    # xx. Crear los objetos ejercicio
    for i in range(len(MODELO_DF)):
        fila_a_revisar = MODELO_DF.iloc[i]
        if not pd.isna(fila_a_revisar['Ejercicios']):
            # Filtro por equipo
            if usar_equipo or fila_a_revisar['Equipo'] == "No":
                ejercicio_actual = ejercicio(fila_a_revisar['Ejercicios'],fila_a_revisar['Nivel'],fila_a_revisar['Equipo'],fila_a_revisar['Página'],fila_a_revisar['Frecuencia'])
                EJERCICIOS_LIST.append(ejercicio_actual)
    
    # xx. Crear los objetos musculo
    nombres_columnas = MODELO_DF.columns.to_list()[5:37]
    for nombre_musculo in nombres_columnas:
        lista_ejercicios_aprobados = []
        for j in range(len(MODELO_DF)):
            supuesta_x = MODELO_DF.iloc[j][nombre_musculo]
            if pd.notna(supuesta_x) and isinstance(supuesta_x, str) and len(supuesta_x) == 1:
                for ejercicio_especifico in EJERCICIOS_LIST:
                    if MODELO_DF.iloc[j]['Ejercicios'] == ejercicio_especifico.nombre:
                        lista_ejercicios_aprobados.append(ejercicio_especifico)
        musculo_actual = musculo(nombre_musculo,lista_ejercicios_aprobados,MODELO_DF.iloc[1][nombre_musculo])
        MUSCULOS_LIST.append(musculo_actual)

    # xx. Crear los objetos grupo muscular
    nombres_musculos = MODELO_DF.columns.to_list()[5:37]
    grupos_ya_verificados = set()
    for nombre in nombres_musculos:
        grupo_en_si = str(MODELO_DF.loc[0,nombre]).split(',')
        frecuencia = MODELO_DF.loc[1,nombre]
        for el in grupo_en_si:
            if el not in grupos_ya_verificados:
                grupos_ya_verificados.add(el)
                objeto_encontrado = None
                for objeto in MUSCULOS_LIST:
                    if objeto.nombre == nombre:
                        objeto_encontrado = objeto
                        break  # Sale del bucle una vez que encuentra el objeto
                grupo_muscular_actual = grupo_muscular(el,[objeto_encontrado],frecuencia)
                GRUPOS_MUSCULARES_LIST.append(grupo_muscular_actual)
            else:
                for objeto in GRUPOS_MUSCULARES_LIST:
                    if objeto.nombre == el:
                        objeto_encontrado = None
                        for objeto_m in MUSCULOS_LIST:
                            if objeto_m.nombre == nombre:
                                objeto_encontrado = objeto_m
                                break
                        objeto.lista_musculos.append(objeto_encontrado)
                        break

    # xx. Crear los objetos tipo Rutina
    rutinas_index = MODELO_DF.columns.to_list()[50:60]
    for id in rutinas_index:
        list_elements_column = list(MODELO_DF[id].dropna().unique())
        list_grupos_musculares = []
        for objeto in GRUPOS_MUSCULARES_LIST:
            if objeto.nombre in list_elements_column:
                list_grupos_musculares.append(objeto)
        rutina_actual = rutina(id,list_grupos_musculares)
        RUTINAS_LIST.append(rutina_actual)

    # Variables
    rutina_de_hoy = int(MODELO_DF['Rutina de hoy'].iloc[0])
    diccionario_de_rutinas_por_valor = {
        0:['A','B','C','D','E','F','G','H'],
        1:['A','B','C','D','E','F','G','H'],
        2:['I','J'],
        3:'Descanso'
    }
    musculos_de_abdomen_excepcion_de_descanso = set() # Sólo los nombres, no todo el objeto
    RUTINAS_FILTRADAS = []
    musculos_en_descanso = MODELO_DF['Musculos en descanso'].dropna().tolist()
    ejercicios_vetados = set()
    valor_principiante = int(MODELO_DF['Principiante'].dropna().iloc[0])
    valor_intermedio = int(MODELO_DF['Intermedio'].dropna().iloc[0])
    valor_avanzado = int(MODELO_DF['Avanzado'].dropna().iloc[0])
    plazas = []
    intensidad_dict = {'Principiante': 1, 'Intermedio': 2, 'Avanzado': 3}
    pares_rutina_frecuencia = dict()
    lista_frecuencias_de_grupos_musculares_df = pd.DataFrame()

    # PREPARAR LA RUTINA DE HOY
    # xx. Sumar una unidad al valor de Rutina de hoy, en caso de ser 2, será una rutina de tren inferior, en caso de
    # ser menor, será una de tren superior, en caso de ser 3, es un día de descanso y se retorna el valor a 0
    valor_rutina_de_hoy = diccionario_de_rutinas_por_valor[rutina_de_hoy]
    if valor_rutina_de_hoy != 'Descanso':
        rutina_de_hoy = rutina_de_hoy+1
        # xx. Filtrar las rutinas con índices válidos
        RUTINAS_FILTRADAS = [rt for rt in RUTINAS_LIST if rt.index in valor_rutina_de_hoy]
        # xx. Recordar que los músculos del abdomen no necesitan periodos de descanso, añadir la variable de excepción
        grupo_muscular_abdomen = next((gm for gm in GRUPOS_MUSCULARES_LIST if gm.nombre == 'Abdomen'), None)
        for mus in grupo_muscular_abdomen.lista_musculos:
            musculos_de_abdomen_excepcion_de_descanso.add(mus.nombre)
        # xx. Revisar que músculos están en descanso y eliminarlos de todas las listas de los grupos musculares 
        for gp in GRUPOS_MUSCULARES_LIST:
            gp.lista_musculos = [
                musculo for musculo in gp.lista_musculos
                if musculo.nombre not in musculos_en_descanso
            ]
        # xx. Revisar que ejercicios están relacionados a los músculos en descanso y eliminarlos de las listas de ejercicios de los músculos restantes
        # Recorremos la lista de músculos extrayendo los nombres de los ejercicios
        for mus in MUSCULOS_LIST:
            if mus.nombre in musculos_en_descanso:
                for ej in mus.lista_ejercicios:
                    ejercicios_vetados.add(ej.nombre)
        # Eliminamos de las listas de músculos todos los ejercicios vetados
        for mus in MUSCULOS_LIST:
            mus.lista_ejercicios = [
                ejer for ejer in mus.lista_ejercicios
                if ejer.nombre not in ejercicios_vetados
            ]
        # ALGORITMO ESPECIAL PARA SEPARAR LAS PLAZAS DE LOS EJERCICIOS Y ESCOGER LOS PROPIOS EJERCICIOS
        # xx. Repetir esto para las 3 plazas de ejercicios
        for plaza in range(3):
            intensidad = ''
            # xx. valor_general = principante+intermedio+avanzado
            valor_general = valor_principiante+valor_intermedio+valor_avanzado
            # xx. Generar un número aleatorio entre 1 y valor_general
            entero_aleatorio = random.randint(1, valor_general)
            # xx. Escoger la intensidad relacionada a la plaza
            if 0 < entero_aleatorio <= valor_principiante:
                intensidad = 'Principiante'
            elif valor_principiante < entero_aleatorio <= valor_principiante+valor_intermedio:
                intensidad = 'Intermedio'
            elif valor_principiante+valor_intermedio < entero_aleatorio <= valor_general:
                intensidad = 'Avanzado'
            # NOTA: Las plazas deben ser listas que contengan su id (momento secuencial en el cual saldrá el ejercicio
            # relacionado), su intensidad y un espacio para poner finalmente el ejercicio que se realizará, inicialmente vacío.
            plazas.append([plaza+1,intensidad,None])
        # LAS PLAZAS YA ESTÁN CREADAS, AHORA A ESCOGER LOS EJERCICIOS
        # xx. Se ordenan las plazas desde la que tiene mayor intensidad hasta la de menor intensidad
        # Ordenar la lista usando la intensidad como clave
        plazas = sorted(plazas, key=lambda x: intensidad_dict[x[1]], reverse=True)
        # xx. A cada rutina le asignamos un valor dependiendo de la suma de las frecuencias con que hayan sido 
        # ejercitados sus grupos musculares. (Para priorizar aquellas cuyos grupos musculares se hayan ejercitado menos)
        for rut in RUTINAS_FILTRADAS:
            g_musculares = [
                gp.nombre for gp in rut.lista_grupos_musculares
            ]
            lista_frecuencias_de_grupos_musculares = MODELO_DF[g_musculares].iloc[0].to_list()
            pares_rutina_frecuencia[rut.index] = sum(lista_frecuencias_de_grupos_musculares)
        # xx. Ordenamos el dict de rutinas desde la de menor valor anterior hasta el mayor
        rutinas_ordenadas = sorted(pares_rutina_frecuencia.items(), key=lambda item: item[1])
        # xx. for rutina in lista_rutinas:
        rutina_escogida = None
        for rutina_ordenada in rutinas_ordenadas:
            rutina_actual = [obj for obj in RUTINAS_FILTRADAS if obj.index == rutina_ordenada[0]][0]
            rutina_escogida = rutina_actual
            # xx. lista_grupo_muscular_actual = rutina.lista_grupos_musculares
            lista_grupo_muscular_actual = rutina_actual.lista_grupos_musculares
            # xx. Verificar si cada grupo muscular de lista_grupo_muscular_actual tiene por lo menos un músculo libre para entrenar
            # En caso de que no, se procede a la siguiente iteración.
            saltar_a_la_siguiente = False
            for gp in lista_grupo_muscular_actual:
                if len(gp.lista_musculos) < 1:
                    saltar_a_la_siguiente = True
                    break
            if saltar_a_la_siguiente:
                continue
            # xx. Se crea una nueva estructura de datos que contiene la siguiente información:
            # Tuplas con nombre del grupo muscular, nombre de músculo, ejercicio e intensidad. 
            # El objetivo es que estas tuplas vayan siendo obtenidas en un orden específico.
            tuplas_gp_mus_ejr_intensidad = []
            # xx. Se ordenan de menor frecuencia a mayor frecuencia todos los grupos musculares presentes en la rutina.
            lista_frecuencias_de_grupos_musculares_df = MODELO_DF[[objeto.nombre for objeto in lista_grupo_muscular_actual]].iloc[0]
            tupla_grupo_muscular_frecuencia_ordenada = sorted(list(lista_frecuencias_de_grupos_musculares_df.items()),key=lambda x: x[1])
            # xx. Para cada uno de estos grupos musculares, se ordena su lista de músculos también de menor a mayor frecuencia.
            for tupla_gp in tupla_grupo_muscular_frecuencia_ordenada:
                objeto_gp_musculos_lista = next(obj for obj in GRUPOS_MUSCULARES_LIST if obj.nombre == tupla_gp[0]).lista_musculos
                lista_frecuencias_de_musculos_df = MODELO_DF[[objeto.nombre for objeto in objeto_gp_musculos_lista]].iloc[1]
                tupla_musculo_frecuencia_ordenada = sorted(list(lista_frecuencias_de_musculos_df.items()),key=lambda x: x[1])
                # xx. Para cada uno de los musculos se ordena de menor a mayor frecuencia sus ejercicios asociados
                for tupla_mus in tupla_musculo_frecuencia_ordenada:
                    objeto_mus_ejercicios_lista = next(obj for obj in MUSCULOS_LIST if obj.nombre == tupla_mus[0]).lista_ejercicios
                    ejercicios_frecuencias_series = MODELO_DF.set_index('Ejercicios')['Frecuencia'].loc[[objeto.nombre for objeto in objeto_mus_ejercicios_lista]]
                    tupla_ejercicio_frecuencia_ordenada = sorted(ejercicios_frecuencias_series.items(),key=lambda x: x[1])
                    # y se crea la tupla con la información y se añade a la lista con un append
                    for tupla_ejr in tupla_ejercicio_frecuencia_ordenada:
                        gp = tupla_gp[0]
                        mus = tupla_mus[0]
                        ejr = tupla_ejr[0]
                        niv = ''
                        for ejr_nivel in objeto_mus_ejercicios_lista:
                            if ejr_nivel.nombre == ejr:
                                niv = ejr_nivel.nivel
                        if niv == '':
                            raise ValueError(f'No hay intensidad definida para el ejercicio {ejr}')
                        tuplas_gp_mus_ejr_intensidad.append((gp,mus,ejr,niv))
            copia_tuplas_gp_mus_ejr_intensidad = tuplas_gp_mus_ejr_intensidad.copy()
            # xx. for plaza in plazas_ordenadas:
            for plaza in plazas:
                # xx. Se busca la primera tupla que cumpla con el requisito de intensidad de la plaza
                tuplas_gp_mus_ejr_intensidad = rellenar_una_plaza(plaza, tuplas_gp_mus_ejr_intensidad)
                # xx. Si no se encuentra, se cambia la intensidad de esa plaza a su intensidad inmediatamente menor 
                # y se realiza nuevamente la búsqueda.
                # Si en la intensidad más baja no se encuentra cómo rellenar la plaza, se salta a verificar la siguiente rutina.
                niveles = {'Avanzado': 'Intermedio', 'Intermedio': 'Principiante', 'Principiante': None}
                while plaza[2] is None:
                    if plaza[1] == 'Principiante':
                        saltar_a_la_siguiente = True
                        break
                    # Reducir la intensidad al siguiente nivel
                    plaza[1] = niveles[plaza[1]]
                    if plaza[1] is None:  # Si ya está en el nivel más bajo, se detiene
                        saltar_a_la_siguiente = True
                        break
                    tuplas_gp_mus_ejr_intensidad = rellenar_una_plaza(plaza, tuplas_gp_mus_ejr_intensidad)
            # xx. Si al final todas las plazas fueron llenadas, se da por concluído el proceso.
            hay_none = any(sublista[-1] is None for sublista in plazas)
            if not hay_none:
                break
            # xx. Si queda alguna plaza por llenar, verificar si es porque todos los grupos musculares ya fueron sacados de las posibilidades.
            # En dado caso, se reinicia el proceso para llenar las plazas faltantes.
            else:
                # Extraer nombres de las plazas (ignorando None)
                nombres_a_eliminar = {
                    plaza[2].nombre for plaza in plazas if plaza[2] is not None
                }
                # Filtrar la lista de tuplas
                copia_tuplas_gp_mus_ejr_intensidad = [
                    tupla for tupla in copia_tuplas_gp_mus_ejr_intensidad
                    if tupla[2] not in nombres_a_eliminar
                ]
                for plaza in plazas:
                    if plaza[2] == None:
                        copia_tuplas_gp_mus_ejr_intensidad = rellenar_una_plaza(plaza, copia_tuplas_gp_mus_ejr_intensidad)
                        niveles = {'Avanzado': 'Intermedio', 'Intermedio': 'Principiante', 'Principiante': None}
                        while plaza[2] is None:
                            if plaza[1] == 'Principiante':
                                saltar_a_la_siguiente = True
                                break
                            # Reducir la intensidad al siguiente nivel
                            plaza[1] = niveles[plaza[1]]
                            if plaza[1] is None:  # Si ya está en el nivel más bajo, se detiene
                                saltar_a_la_siguiente = True
                                break
                            copia_tuplas_gp_mus_ejr_intensidad = rellenar_una_plaza(plaza, copia_tuplas_gp_mus_ejr_intensidad)
            hay_none = any(sublista[-1] is None for sublista in plazas)
            if not hay_none:
                break
            else:
                saltar_a_la_siguiente = True
            if saltar_a_la_siguiente:
                continue
            else:
                break
        # xx. Si el for de rutinas termina y hay plazas que no fueron llenadas, se declara el día como de descanso anticipado.
        hay_none = any(sublista[-1] is None for sublista in plazas)
        if hay_none:
            with open(rutina_txt,'w') as archivo:
                archivo.write('Descanso anticipado')
        else:
            ejercicios_de_hoy = [objeto[2].nombre for objeto in plazas]
            # xx. Si todas las plazas están llenadas, se añaden los músculos a descansar y se actualizan todas las demás variables necesarias.
            musculos_de_hoy = []
            # Cargar el archivo existente
            workbook = load_workbook(MODELO)
            # Seleccionar una hoja de trabajo
            hoja = workbook["Hoja1"]
            # xx. Agregar nuevos músculos que estén en descanso
            # Identificar la última fila con datos en la columna BJ
            columna_BJ = hoja["BJ"]  # Obtiene todas las celdas de la columna BJ
            ultima_fila_BJ = max((celda.row for celda in columna_BJ if celda.value is not None), default=0)
            # xx. Rellenar la lista con los músculos ejercitados hoy
            columna_A = hoja["A"]  # Obtiene todas las celdas de la columna A
            ultima_fila_A = max((celda.row for celda in columna_A if celda.value is not None), default=0)
            for fila in range(1,ultima_fila_A+1):
                if hoja[f"A{fila}"].value in ejercicios_de_hoy:
                    fila_a_evaluar = hoja[fila]
                    ultima_columna = max((celda.column for celda in fila_a_evaluar if celda.value is not None), default=0)
                    for columna in range(1,ultima_columna+1):
                        celda_especifica = hoja.cell(row=fila, column=columna)
                        if celda_especifica.value == 'x' or celda_especifica.value == 'X':
                            musculos_de_hoy.append(hoja.cell(row=1, column=columna).value)
            # Agregar los nuevos valores a la columna BJ
            for i, valor in enumerate(musculos_de_hoy, start=1):
                hoja.cell(row=ultima_fila_BJ + i, column=62, value=valor)  # Columna BJ es la columna 62 (A=1, B=2, ..., BJ=62)
            # xx. Incrementar la frecuencia de los ejercicios
            for ejercicio_de_hoy in ejercicios_de_hoy:
                # Recorrer las filas de la columna A para encontrar el nombre
                for fila in hoja.iter_rows(min_col=1, max_col=1, values_only=False):  # Recorremos la columna A (columna 1)
                    celda_nombre = fila[0]  # Primera columna (columna A)
                    if celda_nombre.value == ejercicio_de_hoy:
                        fila_numero = celda_nombre.row  # Número de la fila donde se encontró el nombre
                        celda_e = hoja.cell(row=fila_numero, column=5)  # Columna E es la columna 5
                        if isinstance(celda_e.value, (int, float)):  # Verificar que sea un número
                            celda_e.value += 1  # Incrementar el valor en 1
                        else:
                            print(f"La celda E{fila_numero} no contiene un número. Se omite.")
                        break  # Salimos del bucle después de encontrar el nombre
                else:
                    print(f"No se encontró el nombre '{ejercicio_de_hoy}' en la columna A.")
            # xx. Incrementar las frecuencias de los músculos y los grupos musculares
            grupos_musculares = set([tupla[0] for tupla in TUPLAS_GP_MUS_EJC_LIST])
            ejercicios = set([tupla[2] for tupla in TUPLAS_GP_MUS_EJC_LIST])
            musculos = set()
            for mus_element in MUSCULOS_LIST:
                lista_de_ejercicios = mus_element.lista_ejercicios
                for ej in lista_de_ejercicios:
                    if ej.nombre in ejercicios:
                        musculos.add(mus_element.nombre)
            # MUSCULOS
            # Iterar por las columnas
            for columna in hoja.iter_cols(min_row=1, max_row=1):  # Iteramos solo la primera fila
                celda = columna[0]  # Primera fila de la columna actual
                if celda.value in musculos:  # Si el nombre está en el set
                    fila_3_celda = hoja.cell(row=3, column=celda.column)  # Obtener celda en fila 3
                    if isinstance(fila_3_celda.value, (int, float)):  # Verificar si tiene un número
                        fila_3_celda.value += 1  # Sumar 1
            # GRUPOS MUSCULARES
            # Iterar por las columnas
            for columna in hoja.iter_cols(min_row=1, max_row=1):  # Iteramos solo la primera fila
                celda = columna[0]  # Primera fila de la columna actual
                if celda.value in grupos_musculares:  # Si el nombre está en el set
                    fila_3_celda = hoja.cell(row=2, column=celda.column)  # Obtener celda en fila 3
                    if isinstance(fila_3_celda.value, (int, float)):  # Verificar si tiene un número
                        fila_3_celda.value += 1  # Sumar 1
            # xx. Incrementar el número de la sesión
            hoja["AO2"] = hoja["AO2"].value + 1
            # xx. Incrementar el número de la rutina
            hoja["BI2"] = rutina_de_hoy
            # xx. Se crea un txt con la rutina.
            with open(rutina_txt,'w') as archivo:
                archivo.write(f'Rutina escogida: {rutina_escogida.index}\n')
                archivo.write('\n')
                for plaza in plazas:
                    archivo.write(f'Ejercicio: {plaza[0]}\n')
                    archivo.write(f'Nombre: {plaza[2].nombre}\n')
                    archivo.write(f'Nivel: {plaza[1]}\n')
                    archivo.write(f'Página: {plaza[2].pagina}\n')
                    archivo.write(f'\n')
            # xx. Se da espacio a esperar que la rutina culmine para dar la retroalimentación:
            # Crear la ventana principal (no se mostrará)
            root = tk.Tk()
            root.withdraw()  # Ocultar la ventana principal
            niveles_presentes = set([objeto[1] for objeto in plazas])
            # Mostrar un cuadro de diálogo de tipo 'Sí/No'
            respuesta = messagebox.askyesno("Alerta", "¿Rutina finalizada exitosamente?")
            if respuesta:
                # xx. Si la rutina fue hecha satisfactoriamente, se realizan las modificaciones necesarias en las probabilidades de las intensidades.
                if 'Avanzado' in niveles_presentes:
                    hoja["AN2"] = hoja["AN2"].value + 1
                elif 'Intermedio' in niveles_presentes:
                    hoja["AN2"] = hoja["AN2"].value + 1
                elif 'Principiante' in niveles_presentes:
                    hoja["AM2"] = hoja["AM2"].value + 1
            else:
                # xx. Si no, se realiza la modificación contraria a las probabilidades para bajar el nivel de la siguiente rutina.
                if 'Avanzado' in niveles_presentes:
                    hoja["AN2"] = hoja["AN2"].value - 1
                elif 'Intermedio' in niveles_presentes:
                    hoja["AN2"] = hoja["AN2"].value - 1
                elif 'Principiante' in niveles_presentes:
                    hoja["AM2"] = hoja["AM2"].value - 1
            # xx. Guardar los cambios
            workbook.save(MODELO)
    # Opción de día de descanso
    else:
        # Cargar el archivo existente
        workbook = load_workbook(MODELO)
        # Seleccionar una hoja de trabajo
        hoja = workbook["Hoja1"]
        # xx. Vaciar la columna de músculos en descanso
        # Recorrer las filas desde la fila inicial hacia abajo
        for fila in range(2, hoja.max_row + 1):
            # Vaciar el contenido de la celda
            hoja[f'BJ{fila}'] = None
        # xx. Actualizar la variable de rutina de hoy
        hoja["BI2"] = 0
        workbook.save(MODELO)
        with open(rutina_txt,'w') as archivo:
            archivo.write('Día de descanso')